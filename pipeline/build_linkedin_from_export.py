#!/usr/bin/env python3
"""
build_linkedin_from_export.py
-----------------------------
Turns the LinkedIn Page admin exports into the dashboard's LinkedIn source of
truth — replacing the fabricated SAMPLE data with real numbers.

Handles LinkedIn's native `.xls` (old OLE2 binary, read via xlrd) as well as
.xlsx/.csv. Understands the standard LinkedIn export sheets:
  followers file -> "New followers" sheet: Date + daily-new counts. The
      "Total followers" column is DAILY NEW (not cumulative), so we running-sum
      it to reconstruct the cumulative follower curve + current total.
  content file   -> "All posts" sheet (per-post table) + "Metrics" sheet
      (daily aggregated impressions/reactions/comments/reposts for the trend).
  visitors file  -> "Visitor metrics" sheet: daily page views + unique visitors.

Writes data/linkedin_latest.json (source:"export") which the LinkedInPanel
reads, and banks linkedin/followers history + linkedin/impressions into
data/snapshots.json (idempotent). Headline summary figures use the last 30 days;
the daily arrays keep the full range so the 7D/30D/6M/1Y tabs all work.

Usage:
  python pipeline/build_linkedin_from_export.py \
     --followers <followers.xls> --content <content.xls> --visitors <visitors.xls>
  # or auto-discover by filename in --dir (default: repo root, uploads, data)
"""

import os
import re
import sys
import csv
import glob
import json
import argparse
from datetime import datetime, timezone, timedelta

from save_snapshot import save_snapshot

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data")
OUT = os.path.join(DATA, "linkedin_latest.json")
ORG_NAME = os.getenv("LINKEDIN_COMPANY", "mangrove-technologies-inc")
ORG_URL = f"https://www.linkedin.com/company/{ORG_NAME}/"
WINDOW = 30


# ---------- generic sheet reading (xls via xlrd, xlsx via openpyxl, csv) ----------
def read_sheet(path, sheet=None, header_row=0):
    """Return (headers, list-of-dict rows) for the given sheet."""
    low = path.lower()
    if low.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    elif low.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            grid = list(csv.reader(f))
    if header_row >= len(grid):
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in grid[header_row]]
    rows = [dict(zip(headers, r)) for r in grid[header_row + 1:] if any(c not in (None, "") for c in r)]
    return headers, rows


def col(headers, *cands):
    low = [h.lower() for h in headers]
    for cand in cands:
        cl = cand.lower()
        for i, h in enumerate(low):
            if h == cl:
                return headers[i]
    for cand in cands:
        cl = cand.lower()
        for i, h in enumerate(low):
            if cl in h:
                return headers[i]
    return None


def num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def pdate(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    v = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else None


def last_n_sum(daily, key, days=WINDOW):
    if not daily:
        return 0
    maxd = max(r["date"] for r in daily)
    cut = (datetime.strptime(maxd, "%Y-%m-%d") - timedelta(days=days)).strftime("%Y-%m-%d")
    return int(sum(r.get(key, 0) for r in daily if r["date"] >= cut))


# ---------- parsers ----------
def parse_followers(path):
    for sheet, hr in (("New followers", 0), (None, 0)):
        headers, rows = read_sheet(path, sheet, hr)
        d = col(headers, "Date"); t = col(headers, "Total followers", "New followers")
        if d and t:
            running, hist = 0, []
            for r in sorted(rows, key=lambda x: pdate(x.get(d)) or ""):
                dt = pdate(r.get(d))
                if not dt:
                    continue
                running += int(num(r.get(t)) or 0)
                hist.append({"date": dt, "followers": running})
            return hist, f"followers: {len(hist)} days via '{t}' (running-sum → cumulative)"
    return [], f"followers: columns not matched"


def parse_content(path):
    posts = []
    ph, prows = read_sheet(path, "All posts", 1)
    if prows:
        cT, cU, cD = col(ph, "Post title"), col(ph, "Post link", "Post url"), col(ph, "Created date", "Date")
        cI, cL, cC, cR, cE = (col(ph, "Impressions"), col(ph, "Likes", "Reactions"),
                              col(ph, "Comments"), col(ph, "Reposts", "Shares"), col(ph, "Engagement rate"))
        for i, r in enumerate(prows):
            imp = num(r.get(cI)) if cI else None
            if imp is None:
                continue
            reac = int(num(r.get(cL)) or 0); com = int(num(r.get(cC)) or 0); sh = int(num(r.get(cR)) or 0)
            er = num(r.get(cE)) if cE else None
            if er is None:
                er = round((reac + com + sh) / imp * 100, 3) if imp else 0.0
            posts.append({"id": f"post{i}", "date": pdate(r.get(cD)) if cD else None,
                          "text": str(r.get(cT) or "").strip()[:200], "impressions": int(imp),
                          "reactions": reac, "comments": com, "shares": sh, "engagement_rate": er,
                          "url": (str(r.get(cU)).strip() if cU and r.get(cU) else None)})
    daily = []
    mh, mrows = read_sheet(path, "Metrics", 1)
    if mrows:
        cD = col(mh, "Date"); cI = col(mh, "Impressions (total)", "Impressions")
        cRe = col(mh, "Reactions (total)"); cCo = col(mh, "Comments (total)"); cRp = col(mh, "Reposts (total)")
        for r in mrows:
            dt = pdate(r.get(cD))
            if not dt:
                continue
            imp = int(num(r.get(cI)) or 0) if cI else 0
            re_ = int(num(r.get(cRe)) or 0) if cRe else 0
            co_ = int(num(r.get(cCo)) or 0) if cCo else 0
            rp_ = int(num(r.get(cRp)) or 0) if cRp else 0
            daily.append({"date": dt, "impressions": imp, "engagements": re_ + co_ + rp_,
                          "reactions": re_, "comments": co_, "shares": rp_})
    daily.sort(key=lambda r: r["date"])
    posts.sort(key=lambda p: p["impressions"], reverse=True)
    return posts, daily, f"content: {len(posts)} posts, {len(daily)} daily rows"


def parse_visitors(path):
    for sheet, hr in (("Visitor metrics", 0), (None, 0)):
        headers, rows = read_sheet(path, sheet, hr)
        cD = col(headers, "Date")
        cPV = col(headers, "Total page views (total)", "Overview page views (total)", "Page views (total)")
        cUV = col(headers, "Total unique visitors (total)", "Overview unique visitors (total)", "Unique visitors (total)")
        if cD and (cPV or cUV):
            per_day = []
            for r in rows:
                dt = pdate(r.get(cD))
                if not dt:
                    continue
                per_day.append({"date": dt, "pv": int(num(r.get(cPV)) or 0) if cPV else 0,
                                "uv": int(num(r.get(cUV)) or 0) if cUV else 0})
            return per_day, f"visitors: {len(per_day)} days (pv='{cPV}', uv='{cUV}')"
    return [], "visitors: columns not matched"


# ---------- discovery ----------
def discover(kind, explicit, dirs):
    if explicit:
        return explicit
    pats = {"followers": ["*ollower*"], "content": ["*ontent*", "*post*"], "visitors": ["*isitor*"]}[kind]
    for d in dirs:
        for pat in pats:
            for ext in ("xls", "xlsx", "csv"):
                hits = sorted(glob.glob(os.path.join(d, f"{pat}.{ext}")))
                if hits:
                    return hits[0]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--followers"); ap.add_argument("--content"); ap.add_argument("--visitors")
    ap.add_argument("--dir")
    a = ap.parse_args()
    dirs = [a.dir] if a.dir else [os.path.join(HERE, ".."), os.path.join(HERE, "..", "..", "uploads"),
            os.path.join(HERE, "..", "uploads"), os.path.join(HERE, "..", "data")]
    dirs = [d for d in dirs if d and os.path.isdir(d)]

    f_file = discover("followers", a.followers, dirs)
    c_file = discover("content", a.content, dirs)
    v_file = discover("visitors", a.visitors, dirs)
    print("Files:", {"followers": f_file, "content": c_file, "visitors": v_file})

    notes, summary = [], {}
    history, posts, daily, visitors = [], [], [], []

    if f_file:
        history, note = parse_followers(f_file); notes.append(note)
    else:
        notes.append("followers: file not found")
    if c_file:
        posts, daily, note = parse_content(c_file); notes.append(note)
    else:
        notes.append("content: file not found")
    if v_file:
        visitors, note = parse_visitors(v_file); notes.append(note)
    else:
        notes.append("visitors: file not found")

    if not history:
        print("\n".join(notes)); sys.exit("\nNo follower data parsed — cannot mark LinkedIn live.")

    summary["followers"] = history[-1]["followers"]
    # 30-day follower gain from the cumulative curve
    cut = (datetime.strptime(history[-1]["date"], "%Y-%m-%d") - timedelta(days=WINDOW)).strftime("%Y-%m-%d")
    past = next((h["followers"] for h in reversed(history) if h["date"] <= cut), history[0]["followers"])
    summary["follower_growth"] = summary["followers"] - past

    if daily:
        # All headline engagement figures are last-30-days, consistent with
        # impressions/visitors (per-post totals still feed the posts table).
        summary["post_impressions"] = last_n_sum(daily, "impressions")
        summary["post_reactions"] = last_n_sum(daily, "reactions")
        summary["post_comments"] = last_n_sum(daily, "comments")
        summary["post_shares"] = last_n_sum(daily, "shares")
        eng30 = summary["post_reactions"] + summary["post_comments"] + summary["post_shares"]
        summary["engagement_rate"] = round(eng30 / summary["post_impressions"] * 100, 2) if summary["post_impressions"] else None
    if visitors:
        maxd = max(r["date"] for r in visitors)
        cutv = (datetime.strptime(maxd, "%Y-%m-%d") - timedelta(days=WINDOW)).strftime("%Y-%m-%d")
        recent = [r for r in visitors if r["date"] >= cutv]
        summary["page_views"] = sum(r["pv"] for r in recent)
        summary["unique_visitors"] = sum(r["uv"] for r in recent)

    output = {
        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "export", "window_days": WINDOW,
        "organization": {"name": ORG_NAME, "url": ORG_URL, "followers": summary["followers"], "logo_url": None},
        "summary": summary, "followers_history": history, "daily": daily, "posts": posts[:20],
    }
    os.makedirs(DATA, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)

    snap = [{"platform": "linkedin", "metric": "followers", "value": h["followers"], "date": h["date"]} for h in history]
    snap += [{"platform": "linkedin", "metric": "impressions", "value": r["impressions"], "date": r["date"]} for r in daily]
    save_snapshot(snap)

    print("\n".join(notes))
    print(f"\nWrote {OUT}")
    print(f"followers(now)={summary['followers']}  (+{summary['follower_growth']}/30d)  "
          f"tracking since {history[0]['date']}")
    print(f"posts={len(posts)}  impressions(30d)={summary.get('post_impressions')}  "
          f"page_views(30d)={summary.get('page_views')}  unique_visitors(30d)={summary.get('unique_visitors')}")


if __name__ == "__main__":
    main()
