#!/usr/bin/env python3
"""
ingest_linkedin_export.py  (Phase 2.3)
--------------------------------------
Ingest LinkedIn Page follower history from the admin analytics export into the
snapshot store. LinkedIn has no retrievable follower-history API, so the owner
exports "Followers" over time from the Page admin → Analytics → Followers UI
(.xlsx or .csv), and this parses it into dated linkedin/followers snapshots.

Idempotent (upsert by platform/metric/day) — safe to re-run after new exports.

Usage:
  python pipeline/ingest_linkedin_export.py path/to/linkedin-followers.xlsx
  python pipeline/ingest_linkedin_export.py path/to/export.csv

It auto-detects a date column and a followers column (handles LinkedIn's
"Total followers" cumulative column, or a "New followers" column which it
accumulates onto today's known total if needed).
"""

import os
import sys
import csv
import re
from datetime import datetime

from save_snapshot import save_snapshot

DATE_HINTS = ("date", "day")
TOTAL_HINTS = ("total follower", "cumulative", "followers")   # prefer cumulative
NEW_HINTS = ("new follower", "gained")


def _parse_date(v):
    v = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%b %d, %Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _num(v):
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def read_rows(path):
    """Return list of dicts (header -> cell) from .csv or .xlsx."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Reading .xlsx needs openpyxl: pip install openpyxl (or export as CSV).")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Find the header row (first row containing a date-ish and a number-ish header).
        header_i = 0
        for i, r in enumerate(rows[:10]):
            joined = " ".join(str(c).lower() for c in r if c)
            if any(h in joined for h in DATE_HINTS):
                header_i = i
                break
        header = [str(c).strip() if c is not None else "" for c in rows[header_i]]
        return [dict(zip(header, r)) for r in rows[header_i + 1:]]
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def pick(headers, hints):
    for h in headers:
        hl = h.lower()
        if any(x in hl for x in hints):
            return h
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python pipeline/ingest_linkedin_export.py <export.xlsx|.csv>")
    path = sys.argv[1]
    if not os.path.isfile(path):
        sys.exit(f"File not found: {path}")

    rows = read_rows(path)
    if not rows:
        sys.exit("No rows found in export.")
    headers = list(rows[0].keys())
    date_col = pick(headers, DATE_HINTS)
    total_col = pick(headers, TOTAL_HINTS)
    new_col = pick(headers, NEW_HINTS)
    if not date_col or not (total_col or new_col):
        sys.exit(f"Couldn't find date + followers columns in: {headers}")

    parsed = []
    for r in rows:
        d = _parse_date(r.get(date_col))
        if not d:
            continue
        if total_col:
            v = _num(r.get(total_col))
            if v is not None:
                parsed.append((d, v))
        else:
            gained = _num(r.get(new_col)) or 0
            parsed.append((d, gained))  # cumulative reconstructed below

    if not parsed:
        sys.exit("No usable date/value rows parsed.")
    parsed.sort()

    if not total_col:
        running = 0
        cum = []
        for d, gained in parsed:
            running += gained
            cum.append((d, running))
        parsed = cum
        print("Note: export had only 'new followers'; stored a running cumulative "
              "(offset from 0 — adjust if you need absolute totals).")

    out = [{"platform": "linkedin", "metric": "followers", "value": v, "date": d} for d, v in parsed]
    n = save_snapshot(out)
    print(f"Ingested {n} linkedin/followers rows ({parsed[0][0]} → {parsed[-1][0]}).")


if __name__ == "__main__":
    main()
